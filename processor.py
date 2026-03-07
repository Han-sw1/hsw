import pandas as pd
import xlrd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import io


# ─── 상수 ───────────────────────────────────────────────

# 지역버스 처리부서 → 탭명 매핑
REGIONAL_TAB_MAP = {
    "[G1]계룡센터":  "대전 B650",
    "[G1]대전센터":  "대전 B650",
    "[G1]세종센터":  "세종 B500",
    "[G1]제주센터":  "제주 B400",
    "[G1]포항센터":  "포항 B800",
    "[G1]상주센터":  "상주,영주,예천 B400",
    "[G1]영주센터":  "상주,영주,예천 B400",
    "[G1]예천센터":  "상주,영주,예천 B400",
    "[G1]안동센터":  "안동 B520D",
    "[G1]김해센터":  "김해 B600",
}

# 탭 출력 순서
TAB_ORDER = [
    "서울 B800", "서울 B700", "서울 B710", "공항 B620",
    "대전 B650", "세종 B500", "제주 B400", "포항 B800",
    "상주,영주,예천 B400", "안동 B520D", "김해 B600",
]

# 단말기구분 키워드 → 탭명
DEVICE_TAB_MAP = {
    "B800": "서울 B800",
    "B700": "서울 B700",
    "B710": "서울 B710",
    "B620": "공항 B620",
}

# 단말기구분 키워드 → 기준파일 컬럼명 (B620은 헤더 정리 후)
DEVICE_CRITERIA_COL = {
    "B800": "B800",
    "B700": "B700",
    "B710": "B710",
    "B620": "B620(공항)",
}

# 탭별 헤더 배경색
TAB_COLORS = {
    "서울 B800":          "CE0E2D",
    "서울 B700":          "CE0E2D",
    "서울 B710":          "CE0E2D",
    "공항 B620":          "1A5276",
    "대전 B650":          "1E8449",
    "세종 B500":          "1E8449",
    "제주 B400":          "1E8449",
    "포항 B800":          "1E8449",
    "상주,영주,예천 B400": "1E8449",
    "안동 B520D":         "1E8449",
    "김해 B600":          "1E8449",
}


# ─── 날짜/주차 유틸 ────────────────────────────────────

def get_week_label(d):
    """전주목~금주수 기준 주차 레이블 (예: '2월4주')"""
    if not isinstance(d, date):
        return ""
    days_to_wed = (2 - d.weekday()) % 7
    week_end_wed = d + timedelta(days=days_to_wed)
    week_start_thu = week_end_wed - timedelta(days=6)
    month = week_end_wed.month
    year = week_end_wed.year
    first_day = date(year, month, 1)
    first_wednesday = first_day + timedelta(days=(2 - first_day.weekday()) % 7)
    first_thursday = first_wednesday - timedelta(days=6)
    week_num = (week_start_thu - first_thursday).days // 7 + 1
    return f"{month}월{week_num}주"


def parse_date_from_str(val):
    """yyyymmddHHMMSS → date"""
    try:
        s = str(int(float(val)))
        return date(int(s[:4]), int(s[4:6]), int(s[6:8]))
    except Exception:
        return None


def parse_date_from_datetime_str(val):
    """'2026-02-19 12:10' 또는 datetime → date"""
    try:
        if isinstance(val, date):
            return val
        s = str(val).strip()
        return pd.to_datetime(s).date()
    except Exception:
        return None


# ─── 파일 로드 ─────────────────────────────────────────

def load_xls_as_df(path):
    """xls 파일 → DataFrame (cp949 인코딩)"""
    wb = xlrd.open_workbook(path, encoding_override='cp949')
    ws = wb.sheet_by_index(0)
    headers = ws.row_values(0)
    rows = [ws.row_values(i) for i in range(1, ws.nrows)]
    return pd.DataFrame(rows, columns=headers)


def load_file_as_df(path):
    """xls/xlsx 자동 로드"""
    if path.lower().endswith('.xlsx'):
        return pd.read_excel(path, dtype=str)
    else:
        return load_xls_as_df(path)


def detect_file_type(df):
    """DataFrame 내용으로 파일 유형 감지"""
    cols = list(df.columns)
    # 지역버스: 'No' 컬럼 + '처리부서' 컬럼 존재
    if "No" in cols and "처리부서" in cols:
        return "지역버스"
    if "단말기구분" not in cols:
        return "unknown"
    types_str = " ".join(df["단말기구분"].astype(str).tolist())
    if "B710" in types_str:
        return "B710"
    if "B620" in types_str:
        return "B620"
    return "서울"  # B700, B800 혼재


# ─── 기준파일 로드 ──────────────────────────────────────

def load_error_criteria(criteria_path):
    """오류처리유형 기준 파일 → {코드값명: row_dict}"""
    df = pd.read_excel(criteria_path, dtype=str)
    df.columns = [str(c).strip().replace('\n', '') for c in df.columns]
    criteria = {}
    for _, row in df.iterrows():
        name = str(row.get("코드값명", "")).strip()
        if not name or name == "nan":
            continue
        criteria[name] = {c: str(v).strip() for c, v in row.items()}
    return criteria


def load_cits_map(cits_path):
    """CITS 기준 파일 → {차량번호: 설치일(date)}"""
    df = pd.read_excel(cits_path, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    cits_map = {}
    for _, row in df.iterrows():
        차번 = str(row.get("차량번호", "")).strip()
        설치일 = row.get("설치일", "")
        if 차번 and 차번 != "nan":
            try:
                cits_map[차번] = pd.to_datetime(설치일).date() if 설치일 and 설치일 != "nan" else None
            except Exception:
                cits_map[차번] = None
    return cits_map


# ─── 비장애 판별 ────────────────────────────────────────

def is_비장애_with_col(오류유형, device_criteria_col, criteria):
    """특정 단말기 컬럼 기준으로 비장애 여부 판별"""
    key = str(오류유형).strip()
    if not key or key == "nan":
        return False
    if key not in criteria:
        return False
    row = criteria[key]
    # 해당 단말기에 'o' 없으면 적용 안됨
    if str(row.get(device_criteria_col, "")).lower() != "o":
        return False
    return row.get("장애여부", "").strip() == "비장애"


def is_비장애_any(오류유형, criteria):
    """단말기 구분 없이 비장애 여부 판별 (지역버스용)"""
    key = str(오류유형).strip()
    if not key or key == "nan":
        return False
    if key not in criteria:
        return False
    return criteria[key].get("장애여부", "").strip() == "비장애"


# ─── 서울 B700/B800 처리 ───────────────────────────────

def process_seoul(df, criteria, cits_map):
    """
    서울 원본 → B700 탭 df, B800 탭 df
    returns: {"서울 B700": df, "서울 B800": df}
    """
    result = {}
    for device_key, tab_name in [("B800", "서울 B800"), ("B700", "서울 B700")]:
        d = df[df["단말기구분"].astype(str).str.contains(device_key, na=False)].copy()
        if d.empty:
            continue

        # 재현 필터
        d = d[d["증상 재현여부"].astype(str).str.strip() == "재현"]

        # 비장애 필터
        crit_col = DEVICE_CRITERIA_COL[device_key]
        def ok(row, cc=crit_col):
            if is_비장애_with_col(row.get("접수오류유형", ""), cc, criteria):
                return False
            if is_비장애_with_col(row.get("현장처리유형", ""), cc, criteria):
                return False
            return True
        if not d.empty:
            d = d[d.apply(ok, axis=1).astype(bool)]
        if d.empty:
            continue

        # 날짜 파싱 및 정렬
        d["_날짜"] = d["장애접수일시"].apply(parse_date_from_str)
        d = d.sort_values("_날짜").reset_index(drop=True)

        # 추가 컬럼
        d["날짜"] = d["_날짜"]
        d["cits"] = d.apply(lambda r: cits_map.get(str(r["차량번호"]).strip(), "#N/A"), axis=1)
        d["월"] = d["_날짜"].apply(lambda x: f"{x.month}월" if x else "")
        d["주차"] = d["_날짜"].apply(get_week_label)

        # 컬럼 순서: 장애접수일시 뒤에 날짜/cits/월/주차 삽입
        orig = [c for c in d.columns if c not in ("_날짜", "날짜", "cits", "월", "주차")]
        idx = orig.index("장애접수일시") + 1
        final_cols = orig[:idx] + ["날짜", "cits", "월", "주차"] + orig[idx:]
        result[tab_name] = d[final_cols]

    return result


# ─── B710 / B620 처리 ──────────────────────────────────

def process_b710_b620(df, device_key, tab_name, criteria):
    """
    B710 또는 B620 원본 → 탭 df
    device_key: 'B710' 또는 'B620'
    """
    d = df[df["단말기구분"].astype(str).str.contains(device_key, na=False)].copy()
    if d.empty:
        return None

    # 재현 필터
    d = d[d["증상 재현여부"].astype(str).str.strip() == "재현"]

    # 비장애 필터
    crit_col = DEVICE_CRITERIA_COL[device_key]
    def ok(row, cc=crit_col):
        if is_비장애_with_col(row.get("접수오류유형", ""), cc, criteria):
            return False
        if is_비장애_with_col(row.get("현장처리유형", ""), cc, criteria):
            return False
        return True
    if not d.empty:
        d = d[d.apply(ok, axis=1).astype(bool)]
    if d.empty:
        return None

    # 날짜 파싱 및 정렬
    d["_날짜"] = d["장애접수일시"].apply(parse_date_from_str)
    d = d.sort_values("_날짜").reset_index(drop=True)

    # 추가 컬럼 (CITS 없음)
    d["날짜"] = d["_날짜"]
    d["월"] = d["_날짜"].apply(lambda x: f"{x.month}월" if x else "")
    d["주차"] = d["_날짜"].apply(get_week_label)

    orig = [c for c in d.columns if c not in ("_날짜", "날짜", "월", "주차")]
    idx = orig.index("장애접수일시") + 1
    final_cols = orig[:idx] + ["날짜", "월", "주차"] + orig[idx:]
    return d[final_cols]


# ─── 지역버스 처리 ─────────────────────────────────────

def process_regional(df, criteria):
    """
    지역버스 원본 → {탭명: df} dict
    """
    # No열 삭제
    if "No" in df.columns:
        df = df.drop(columns=["No"])

    # 장애접수일시 → 날짜만 (rename to 장애접수일)
    df["장애접수일시"] = df["장애접수일시"].apply(
        lambda v: parse_date_from_datetime_str(v) if pd.notna(v) else None
    )
    df = df.rename(columns={"장애접수일시": "장애접수일"})

    # 비장애 필터 (단말기현장처리 기준)
    def ok(row):
        return not is_비장애_any(row.get("단말기현장처리", ""), criteria)
    if not df.empty:
        df = df[df.apply(ok, axis=1).astype(bool)]

    # 날짜 정렬
    df = df.sort_values("장애접수일").reset_index(drop=True)

    # 월/주차 추가
    df["월"] = df["장애접수일"].apply(lambda x: f"{x.month}월" if isinstance(x, date) else "")
    df["주차"] = df["장애접수일"].apply(get_week_label)

    # 컬럼 순서: 접수번호, 장애접수일, 월, 주차, 나머지
    other_cols = [c for c in df.columns if c not in ("접수번호", "장애접수일", "월", "주차")]
    final_cols = ["접수번호", "장애접수일", "월", "주차"] + other_cols
    df = df[[c for c in final_cols if c in df.columns]]

    # 처리부서 기준으로 탭 분류
    result = {}
    for _, row in df.iterrows():
        부서 = str(row.get("처리부서", "")).strip()
        tab = REGIONAL_TAB_MAP.get(부서)
        if not tab:
            continue
        if tab not in result:
            result[tab] = []
        result[tab].append(row)

    return {tab: pd.DataFrame(rows).reset_index(drop=True)
            for tab, rows in result.items()}


# ─── 통합 처리 ─────────────────────────────────────────

def process_all_files(file_paths, criteria_path, cits_path):
    """
    여러 파일 경로 → {탭명: df} dict + 메타 정보
    """
    criteria = load_error_criteria(criteria_path)
    cits_map = load_cits_map(cits_path)

    all_tabs = {}   # 탭명 → df 리스트 (누적)
    meta = {}

    for path in file_paths:
        try:
            df = load_file_as_df(path)
        except Exception as e:
            meta[path] = {"error": str(e)}
            continue

        file_type = detect_file_type(df)
        meta[path] = {"type": file_type, "원본행수": len(df)}

        if file_type == "서울":
            tabs = process_seoul(df, criteria, cits_map)
            for tab, tdf in tabs.items():
                all_tabs.setdefault(tab, []).append(tdf)
                meta[path][tab] = len(tdf)

        elif file_type in ("B710", "B620"):
            device_key = file_type
            tab_name = DEVICE_TAB_MAP[device_key]
            tdf = process_b710_b620(df, device_key, tab_name, criteria)
            if tdf is not None and not tdf.empty:
                all_tabs.setdefault(tab_name, []).append(tdf)
                meta[path][tab_name] = len(tdf)

        elif file_type == "지역버스":
            tabs = process_regional(df, criteria)
            for tab, tdf in tabs.items():
                all_tabs.setdefault(tab, []).append(tdf)
                meta[path][tab] = len(tdf)

        else:
            meta[path]["error"] = "파일 유형을 인식할 수 없습니다."

    # 탭별 df 합치기
    final_tabs = {}
    for tab, dfs in all_tabs.items():
        combined = pd.concat(dfs, ignore_index=True)
        # 날짜 기준 재정렬
        date_col = "날짜" if "날짜" in combined.columns else "장애접수일"
        if date_col in combined.columns:
            combined = combined.sort_values(date_col).reset_index(drop=True)
        final_tabs[tab] = combined

    # 집계 메타
    summary = {tab: len(df) for tab, df in final_tabs.items()}

    return final_tabs, summary, meta


# ─── Excel 출력 ────────────────────────────────────────

def make_cell_style(header_color="CE0E2D"):
    thin = Side(style="thin", color="DDDDDD")
    return {
        "header_fill": PatternFill("solid", fgColor=header_color),
        "header_font": Font(bold=True, color="FFFFFF", size=10),
        "header_align": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "odd_fill": PatternFill("solid", fgColor="FFF5F5"),
        "even_fill": PatternFill("solid", fgColor="FFFFFF"),
        "highlight_fill": PatternFill("solid", fgColor="FFE8E8"),
        "cell_font": Font(size=9),
        "cell_align": Alignment(horizontal="center", vertical="center"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


HIGHLIGHT_COLS = {"날짜", "cits", "월", "주차", "장애접수일"}


def write_sheet(ws, df, tab_name):
    style = make_cell_style(TAB_COLORS.get(tab_name, "CE0E2D"))
    headers = list(df.columns)

    # 헤더 행
    for ci, col in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill = style["header_fill"]
        cell.font = style["header_font"]
        cell.alignment = style["header_align"]
        cell.border = style["border"]
    ws.row_dimensions[1].height = 28

    # 데이터 행
    for ri, (_, row) in enumerate(df.iterrows(), 2):
        for ci, col in enumerate(headers, 1):
            val = row[col]
            if val is None or (isinstance(val, float) and pd.isna(val)):
                val = ""
            cell = ws.cell(row=ri, column=ci, value=val)
            if col in HIGHLIGHT_COLS:
                cell.fill = style["highlight_fill"]
            elif ri % 2 == 0:
                cell.fill = style["odd_fill"]
            else:
                cell.fill = style["even_fill"]
            cell.font = style["cell_font"]
            cell.alignment = style["cell_align"]
            cell.border = style["border"]

    # 열 너비
    for ci, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = min(len(str(col)) + 4, 28)

    ws.freeze_panes = "A2"


def tabs_to_excel_bytes(final_tabs):
    """탭별 df dict → 멀티시트 xlsx bytes"""
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    # TAB_ORDER 순서대로 시트 생성
    for tab_name in TAB_ORDER:
        if tab_name not in final_tabs:
            continue
        df = final_tabs[tab_name]
        ws = wb.create_sheet(title=tab_name)
        write_sheet(ws, df, tab_name)

    # TAB_ORDER에 없는 탭 추가 (혹시 있을 경우)
    for tab_name, df in final_tabs.items():
        if tab_name not in TAB_ORDER:
            ws = wb.create_sheet(title=tab_name)
            write_sheet(ws, df, tab_name)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
