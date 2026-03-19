import gc
import os
import re
import zipfile
import shutil
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import xml.sax.saxutils as saxutils
from datetime import date as date_type, datetime

_BASE_DIR = os.environ.get("DATA_DIR", os.path.dirname(__file__))
MONTHLY_FILES_DIR = os.path.join(_BASE_DIR, "monthly_files")

# 처리기 탭명 → 월간 파일 숨김 시트명
TAB_TO_RAWSHEET = {
    "서울 B710":           "B710 로우데이터",
    "서울 B800":           "B800로우데이터",
    "서울 B700":           "B700로우데이터",
    "공항 B620":           "공항 B620로우데이터",
    "대전 B650":           "대전 B650 로우데이터",
    "세종 B500":           "세종B500로우데이터",
    "제주 B400":           "제주 B400 로우데이터",
    "포항 B800":           "포항B800로우데이터",
    "상주,영주,예천 B400": "상주.영주.예천 로우데이터",
    "안동 B520D":          "안동B520D 로우데이터",
    "김해 B600":           "김해B600 로우데이터",
}


def list_monthly_files():
    os.makedirs(MONTHLY_FILES_DIR, exist_ok=True)
    return sorted(fn for fn in os.listdir(MONTHLY_FILES_DIR) if fn.endswith("일자별장애현황.xlsx"))


def get_monthly_file_path(filename):
    return os.path.join(MONTHLY_FILES_DIR, filename)


def _read_sheet(ws):
    """시트 전체를 (헤더 리스트, 데이터 행 리스트)로 읽기. 완전 빈 행 제외."""
    all_rows = list(ws.values)
    if not all_rows:
        return [], []
    headers = list(all_rows[0])
    data = [list(r) for r in all_rows[1:] if any(v is not None for v in r)]
    return headers, data


def _build_col_map(sheet_headers, proc_cols):
    name_to_pos = {}
    for i, h in enumerate(sheet_headers):
        if h is None:
            continue
        hs = str(h).strip()
        name_to_pos.setdefault(hs, []).append(i)

    col_map = {}
    used = set()

    for col in proc_cols:
        col_s = str(col).strip()

        if col_s == "날짜":
            positions = name_to_pos.get("장애접수일시", [])
            available = [p for p in positions if p not in used]
            if len(available) >= 2:
                target = available[1]
            elif available:
                target = available[0]
            else:
                continue
            col_map[col] = target
            used.add(target)

        elif col_s == "cits":
            for cname in ("CITS 설치일", "CITS설치일"):
                positions = name_to_pos.get(cname, [])
                avail = [p for p in positions if p not in used]
                if avail:
                    col_map[col] = avail[0]
                    used.add(avail[0])
                    break

        else:
            positions = name_to_pos.get(col_s, [])
            avail = [p for p in positions if p not in used]
            if avail:
                col_map[col] = avail[0]
                used.add(avail[0])

    return col_map


def _find_col_pos(sheet_headers, *names):
    for name in names:
        for i, h in enumerate(sheet_headers):
            if h is not None and str(h).strip() == name:
                return i
    return None


def _find_date_pos(sheet_headers):
    pos = _find_col_pos(sheet_headers, "장애접수일")
    if pos is not None:
        return pos
    occ = [i for i, h in enumerate(sheet_headers) if h == "장애접수일시"]
    if len(occ) >= 2:
        return occ[1]
    return None


def _prepare_tab_data(final_tabs, mode, week_label):
    prepared = {}
    for tab_name, new_df in final_tabs.items():
        if new_df is None or new_df.empty:
            prepared[tab_name] = {"skipped": True, "reason": "데이터 없음"}
            continue
        sheet_name = TAB_TO_RAWSHEET.get(tab_name)
        if not sheet_name:
            prepared[tab_name] = {"skipped": True, "reason": "시트 매핑 없음"}
            continue
        df_cols = list(new_df.columns)
        df_values = new_df.values.tolist()
        prepared[tab_name] = {
            "skipped": False,
            "sheet_name": sheet_name,
            "df_cols": df_cols,
            "df_values": df_values,
        }
    return prepared


# ─── zipfile 기반 시트 교체 (서식 완전 보존) ───────────────────────────────

def _get_sheet_xml_map(file_path):
    """xlsx에서 시트명 → XML 경로 매핑 반환"""
    ns_wb = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    ns_r = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    with zipfile.ZipFile(file_path, 'r') as zf:
        wb_tree = ET.fromstring(zf.read('xl/workbook.xml'))
        rels_tree = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))
        rid_to_target = {rel.get('Id'): rel.get('Target') for rel in rels_tree}
        sheet_map = {}
        for sheet in wb_tree.findall(f'{{{ns_wb}}}sheets/{{{ns_wb}}}sheet'):
            name = sheet.get('name')
            rid = sheet.get(f'{{{ns_r}}}id')
            target = rid_to_target.get(rid, '')
            if target.startswith('/'):
                sheet_map[name] = target.lstrip('/')
            else:
                sheet_map[name] = f'xl/{target}'
    return sheet_map


def _normalize_date_val(val):
    """날짜값에서 시간(00:00:00) 제거 → 'YYYY-MM-DD' 문자열 반환."""
    if val is None:
        return val
    if isinstance(val, datetime):
        if val.hour == 0 and val.minute == 0 and val.second == 0:
            return val.strftime("%Y-%m-%d")
        return str(val)
    if isinstance(val, date_type):
        return val.strftime("%Y-%m-%d")
    s = str(val)
    if re.match(r'\d{4}-\d{2}-\d{2} 00:00:00$', s):
        return s[:10]
    return val


def _val_to_xml(ref, val):
    """셀 값 → XML 문자열 (날짜 시간 00:00:00 자동 제거)"""
    if val is None:
        return ''
    if isinstance(val, bool):
        return f'<c r="{ref}" t="b"><v>{1 if val else 0}</v></c>'
    if isinstance(val, (int, float)):
        if pd.isna(val):
            return ''
        return f'<c r="{ref}"><v>{val}</v></c>'
    if isinstance(val, (date_type, datetime)):
        s = _normalize_date_val(val)
        safe = saxutils.escape(str(s))
        return f'<c r="{ref}" t="inlineStr"><is><t>{safe}</t></is></c>'
    s = str(val)
    # 문자열 날짜에서 "00:00:00" 제거
    if re.match(r'\d{4}-\d{2}-\d{2} 00:00:00$', s):
        s = s[:10]
    safe = saxutils.escape(s)
    return f'<c r="{ref}" t="inlineStr"><is><t>{safe}</t></is></c>'


def _calc_col_widths(headers, rows, min_w=8, max_w=30):
    """헤더 + 데이터 기준 열 너비 계산"""
    widths = [len(str(h)) if h is not None else 0 for h in headers]
    for row in rows[:2000]:
        for ci, val in enumerate(row):
            if val is not None and ci < len(widths):
                widths[ci] = max(widths[ci], len(str(val)))
    return [max(min_w, min(w + 2, max_w)) for w in widths]


def _make_cols_xml(widths):
    """열 너비 → <cols>...</cols> XML"""
    if not widths:
        return ''
    parts = ['<cols>']
    for i, w in enumerate(widths, 1):
        parts.append(f'<col min="{i}" max="{i}" width="{w}" customWidth="1"/>')
    parts.append('</cols>')
    return ''.join(parts)


def _make_sheetdata_xml(headers, rows):
    """headers + rows → <sheetData>...</sheetData> XML"""
    parts = ['<sheetData>']
    all_rows = [headers] + rows if headers else rows
    for ri, row_data in enumerate(all_rows, 1):
        if not any(v is not None for v in row_data):
            continue
        parts.append(f'<row r="{ri}">')
        for ci, val in enumerate(row_data, 1):
            ref = f'{get_column_letter(ci)}{ri}'
            parts.append(_val_to_xml(ref, val))
        parts.append('</row>')
    parts.append('</sheetData>')
    return ''.join(parts)


def _replace_sheetdata(sheet_xml_bytes, new_sheetdata_xml, new_cols_xml='', total_rows=0, total_cols=0):
    """시트 XML에서 <sheetData> (및 <cols>, <dimension>) 교체"""
    sheet_str = sheet_xml_bytes.decode('utf-8')
    # dimension 업데이트 (openpyxl이 이 범위로 읽을 행 수를 결정함)
    if total_rows > 0 and total_cols > 0:
        new_dim = f'<dimension ref="A1:{get_column_letter(total_cols)}{total_rows}"/>'
        sheet_str = re.sub(r'<dimension[^/]*/>', new_dim, sheet_str)
    # cols 교체 또는 삽입
    if new_cols_xml:
        if re.search(r'<cols[ />]', sheet_str):
            sheet_str = re.sub(r'<cols[^>]*/?>(?:.*?</cols>)?', new_cols_xml, sheet_str, flags=re.DOTALL)
        else:
            sheet_str = sheet_str.replace('<sheetData', new_cols_xml + '<sheetData', 1)
    # sheetData 교체
    result = re.sub(
        r'<sheetData[^>]*/?>(?:.*?</sheetData>)?',
        new_sheetdata_xml,
        sheet_str,
        flags=re.DOTALL,
    )
    return result.encode('utf-8')


def _set_full_calc_on_load(workbook_xml_bytes):
    """workbook.xml의 calcPr에 fullCalcOnLoad="1" 설정 — 열릴 때 수식 강제 재계산."""
    xml = workbook_xml_bytes.decode('utf-8')
    if 'fullCalcOnLoad=' in xml:
        xml = re.sub(r'fullCalcOnLoad="[^"]*"', 'fullCalcOnLoad="1"', xml)
    elif '<calcPr' in xml:
        xml = re.sub(r'(<calcPr\b)', r'\1 fullCalcOnLoad="1"', xml, count=1)
    else:
        xml = xml.replace('</workbook>', '<calcPr fullCalcOnLoad="1"/></workbook>')
    return xml.encode('utf-8')


def _update_xlsx_sheets(file_path, updates):
    """
    xlsx에서 특정 시트의 sheetData만 교체 — 나머지 모든 서식/수식 완전 보존
    updates: {sheet_name: (headers, rows)}
    """
    import io as _io
    sheet_xml_map = _get_sheet_xml_map(file_path)

    # 매핑 안 된 시트 경고
    missing = [sn for sn in updates if sn not in sheet_xml_map]
    if missing:
        print(f"[WARN] _update_xlsx_sheets: 시트 매핑 없음 → {missing}")
        print(f"[WARN] 사용 가능한 시트: {list(sheet_xml_map.keys())}")

    paths_to_update = {
        sheet_xml_map[sn]: (h, r)
        for sn, (h, r) in updates.items()
        if sn in sheet_xml_map
    }
    if not paths_to_update:
        raise RuntimeError(f"수정할 시트를 찾지 못했습니다. updates={list(updates.keys())}, xml_map={list(sheet_xml_map.keys())}")

    buf = _io.BytesIO()
    with zipfile.ZipFile(file_path, 'r') as zf_in:
        with zipfile.ZipFile(buf, 'w', compression=zipfile.ZIP_DEFLATED) as zf_out:
            matched = set()
            for item in zf_in.infolist():
                data = zf_in.read(item.filename)
                if item.filename in paths_to_update:
                    headers, rows = paths_to_update[item.filename]
                    new_sd = _make_sheetdata_xml(headers, rows)
                    widths = _calc_col_widths(headers, rows)
                    new_cols = _make_cols_xml(widths)
                    total_rows = len(rows) + 1  # +1 헤더
                    total_cols = len(headers)
                    data = _replace_sheetdata(data, new_sd, new_cols, total_rows, total_cols)
                    matched.add(item.filename)
                elif item.filename == 'xl/workbook.xml':
                    # 파일 열릴 때 수식 강제 재계산 설정
                    data = _set_full_calc_on_load(data)
                zf_out.writestr(item, data)
        if not matched:
            raise RuntimeError(f"zip 내 파일 경로 불일치. paths_to_update={list(paths_to_update.keys())}")

    with open(file_path, 'wb') as f:
        f.write(buf.getvalue())


# ─── 메인 삽입 함수 ────────────────────────────────────────────────────────

def insert_into_monthly(final_tabs, monthly_filename, mode="daily", week_label=None):
    """
    처리된 데이터를 월간 파일 숨김 시트에 삽입.
    zipfile 방식으로 원본 서식/수식 완전 보존.
    """
    file_path = get_monthly_file_path(monthly_filename)
    if not os.path.exists(file_path):
        return {"error": f"파일 없음: {monthly_filename}"}

    prepared = _prepare_tab_data(final_tabs, mode, week_label)
    del final_tabs
    gc.collect()

    # 현재 raw data 읽기 (read_only, 저장 없음)
    wb = load_workbook(file_path, read_only=True, data_only=True)
    report = {}
    updates = {}  # sheet_name → (headers, merged_rows)

    for tab_name, prep in prepared.items():
        if prep.get("skipped"):
            report[tab_name] = {"skipped": True, "reason": prep.get("reason", "")}
            continue

        sheet_name = prep["sheet_name"]
        if sheet_name not in wb.sheetnames:
            report[tab_name] = {"skipped": True, "reason": f"시트 없음: {sheet_name}"}
            continue

        ws = wb[sheet_name]
        sheet_headers, existing_rows = _read_sheet(ws)

        if not sheet_headers:
            sheet_headers = prep["df_cols"]
            existing_rows = []

        num_cols = len(sheet_headers)
        df_cols = prep["df_cols"]
        df_values = prep["df_values"]

        col_map = _build_col_map(sheet_headers, df_cols)
        new_rows = []
        for row_vals in df_values:
            r = [None] * num_cols
            for ci, col in enumerate(df_cols):
                pos = col_map.get(col)
                if pos is not None and pos < num_cols:
                    val = row_vals[ci]
                    if val is None or (isinstance(val, float) and pd.isna(val)):
                        val = None
                    r[pos] = val
            new_rows.append(r)

        id_pos = _find_col_pos(sheet_headers, "접수번호", "No")
        week_pos = _find_col_pos(sheet_headers, "주차")
        date_pos = _find_date_pos(sheet_headers)

        # 기존 행 정리: 날짜 컬럼 00:00:00 제거 + 접수번호 중복 행 제거
        if date_pos is not None:
            for row in existing_rows:
                if len(row) > date_pos:
                    row[date_pos] = _normalize_date_val(row[date_pos])
        if id_pos is not None:
            seen_ids: set = set()
            cleaned = []
            for row in existing_rows:
                rid = str(row[id_pos]) if len(row) > id_pos and row[id_pos] is not None else None
                if rid is None or rid not in seen_ids:
                    if rid:
                        seen_ids.add(rid)
                    cleaned.append(row)
            existing_rows = cleaned

        before = len(existing_rows)

        if mode == "monthly":
            merged = new_rows
        elif mode == "weekly" and week_label:
            if week_pos is not None:
                # 주차 컬럼 있으면: 해당 주차 기존 데이터 삭제 후 교체
                keep = [r for r in existing_rows if str(r[week_pos]) != str(week_label)]
            else:
                # 주차 컬럼 없으면(지역 탭 등): 접수번호 기준 중복 제거
                if id_pos is not None:
                    id_col_idx = next((ci for ci, c in enumerate(df_cols) if col_map.get(c) == id_pos), None)
                    if id_col_idx is not None:
                        new_ids = {str(vals[id_col_idx]) for vals in df_values}
                        keep = [r for r in existing_rows if not (len(r) > id_pos and r[id_pos] is not None and str(r[id_pos]) in new_ids)]
                    else:
                        keep = existing_rows
                else:
                    keep = existing_rows
            merged = keep + new_rows
        else:
            if id_pos is not None:
                existing_ids = {str(r[id_pos]) for r in existing_rows if r[id_pos] is not None}
                id_col_idx = next((ci for ci, c in enumerate(df_cols) if col_map.get(c) == id_pos), None)
                if id_col_idx is not None:
                    to_add = [r for r, vals in zip(new_rows, df_values) if str(vals[id_col_idx]) not in existing_ids]
                else:
                    to_add = new_rows
                merged = existing_rows + to_add
            else:
                existing_tuples = {tuple(str(v) if v is not None else "" for v in r) for r in existing_rows}
                to_add = [r for r in new_rows if tuple(str(v) if v is not None else "" for v in r) not in existing_tuples]
                merged = existing_rows + to_add

        if date_pos is not None and merged:
            def _date_sort_key(r):
                val = r[date_pos]
                if val is None:
                    return (1, '')
                if isinstance(val, (date_type, datetime)):
                    return (0, val.isoformat())
                try:
                    return (0, pd.to_datetime(str(val)).date().isoformat())
                except Exception:
                    return (0, str(val))
            try:
                merged.sort(key=_date_sort_key)
            except Exception:
                pass

        updates[sheet_name] = (sheet_headers, merged)
        report[tab_name] = {
            "sheet": sheet_name,
            "before": before,
            "added": len(merged) - before,
            "total": len(merged),
        }

    wb.close()
    del wb
    gc.collect()

    if updates:
        _update_xlsx_sheets(file_path, updates)

    # DB 통계 갱신
    try:
        import database as db
        from analytics import read_monthly_stats, _parse_ym
        year, month = _parse_ym(monthly_filename)
        if year:
            stats = read_monthly_stats(monthly_filename)
            if stats:
                db.upsert_file_stats(monthly_filename, year, month, stats)
    except Exception as e:
        print(f"[DB] 통계 갱신 오류 ({monthly_filename}): {e}")

    return report


def cleanup_monthly_duplicates(monthly_filename):
    """월간 파일 모든 로우데이터 시트에서 날짜 정규화 + 접수번호 중복 제거."""
    file_path = get_monthly_file_path(monthly_filename)
    if not os.path.exists(file_path):
        return {"error": f"파일 없음: {monthly_filename}"}

    wb = load_workbook(file_path, read_only=True, data_only=True)
    updates = {}
    report = {}

    for tab_name, sheet_name in TAB_TO_RAWSHEET.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        sheet_headers, existing_rows = _read_sheet(ws)
        if not sheet_headers or not existing_rows:
            continue

        id_pos = _find_col_pos(sheet_headers, "접수번호", "No")
        date_pos = _find_date_pos(sheet_headers)
        before = len(existing_rows)

        # 날짜 정규화 (00:00:00 제거)
        if date_pos is not None:
            for row in existing_rows:
                if len(row) > date_pos:
                    row[date_pos] = _normalize_date_val(row[date_pos])

        # 접수번호 기준 중복 제거
        if id_pos is not None:
            seen_ids: set = set()
            cleaned = []
            for row in existing_rows:
                rid = str(row[id_pos]) if len(row) > id_pos and row[id_pos] is not None else None
                if rid is None or rid not in seen_ids:
                    if rid:
                        seen_ids.add(rid)
                    cleaned.append(row)
            existing_rows = cleaned

        after = len(existing_rows)
        updates[sheet_name] = (sheet_headers, existing_rows)
        if before != after:
            report[tab_name] = {"before": before, "after": after, "removed": before - after}

    wb.close()
    del wb
    gc.collect()

    if updates:
        _update_xlsx_sheets(file_path, updates)

    return report


def delete_weekly(monthly_filename, week_label):
    """지정 주차 데이터를 월간 파일 모든 시트에서 삭제."""
    file_path = get_monthly_file_path(monthly_filename)
    if not os.path.exists(file_path):
        return {"error": f"파일 없음: {monthly_filename}"}

    wb = load_workbook(file_path, read_only=True, data_only=True)
    report = {}
    updates = {}

    for tab_name, sheet_name in TAB_TO_RAWSHEET.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers, rows = _read_sheet(ws)
        if not headers:
            continue
        week_pos = _find_col_pos(headers, "주차")
        if week_pos is None:
            continue
        before = len(rows)
        kept = [r for r in rows if str(r[week_pos]) != str(week_label)]
        deleted = before - len(kept)
        if deleted > 0:
            updates[sheet_name] = (headers, kept)
        report[tab_name] = {"sheet": sheet_name, "deleted": deleted, "remaining": len(kept)}

    wb.close()
    del wb
    gc.collect()

    if updates:
        _update_xlsx_sheets(file_path, updates)

    return report


def _get_date_series(df):
    """날짜 컬럼을 date 객체 시리즈로 반환."""
    date_col = "날짜" if "날짜" in df.columns else "장애접수일"
    if date_col not in df.columns:
        return None
    def to_date(v):
        if isinstance(v, date_type):
            return v
        try:
            return pd.to_datetime(str(v)).date()
        except Exception:
            return None
    return df[date_col].apply(to_date)


def compute_web_stats(final_tabs):
    """
    처리된 데이터에서 웹 통계 계산.
    """
    stats = {}
    for tab_name, df in final_tabs.items():
        if df is None or df.empty:
            continue

        from analytics import REGIONAL_TABS
        is_regional = tab_name in REGIONAL_TABS
        fault_col = "단말기접수유형" if is_regional else "접수오류유형"

        date_series = _get_date_series(df)

        by_month = {}
        primary_month = None
        if date_series is not None:
            month_vals = date_series.apply(lambda d: f"{d.month}월" if d else None).dropna()
            if not month_vals.empty:
                by_month = {str(k): int(v) for k, v in month_vals.value_counts().sort_index().items()}
                # 주차 레이블에서 월 추출 (예: "3월1주" → 3월)
                if "주차" in df.columns:
                    import re as _re
                    def _week_month(w):
                        m = _re.search(r'(\d+)월', str(w))
                        return int(m.group(1)) if m else None
                    week_months = df["주차"].dropna().apply(_week_month).dropna()
                    if not week_months.empty:
                        primary_month = f"{int(week_months.mode()[0])}월"
                if primary_month is None:
                    primary_month = month_vals.value_counts().idxmax()

        total_primary = 0
        if primary_month and date_series is not None:
            pm_num = int(primary_month.replace("월", ""))
            mask = date_series.apply(lambda d: d is not None and d.month == pm_num)
            total_primary = int(mask.sum())
        else:
            total_primary = len(df)

        top3 = ""
        if fault_col in df.columns:
            if primary_month and date_series is not None:
                pm_num = int(primary_month.replace("월", ""))
                mask = date_series.apply(lambda d: d is not None and d.month == pm_num)
                vals = df.loc[mask, fault_col].astype(str).str.strip()
            else:
                vals = df[fault_col].astype(str).str.strip()
            vals = vals[~vals.isin(["", "nan", "None", "NaN"])]
            if not vals.empty:
                top = vals.value_counts().head(3)
                top3 = "/".join(top.index.tolist())

        by_week = {}
        if "주차" in df.columns:
            by_week = {str(k): int(v) for k, v in df["주차"].value_counts().sort_index().items()}

        weeks = sorted(by_week.keys())

        by_date = {}
        if date_series is not None:
            def _to_d_str(d):
                if d is None:
                    return None
                try:
                    return f"{d.month}/{d.day}"
                except Exception:
                    return None
            date_strs = date_series.apply(_to_d_str).dropna()
            by_date = {str(k): int(v) for k, v in date_strs.value_counts().sort_index().items()}

        # 주차별 오류유형 건수
        fault_by_week = {}
        if "주차" in df.columns and fault_col in df.columns:
            for wk_val, wk_group in df.groupby("주차"):
                fault_cnt = {}
                for ft, cnt in wk_group[fault_col].value_counts().items():
                    ft_str = str(ft).strip()
                    if ft_str and ft_str not in ('', 'nan', 'None', 'NaN'):
                        fault_cnt[ft_str] = int(cnt)
                if fault_cnt:
                    fault_by_week[str(wk_val)] = fault_cnt

        # 전체(주 대상월) 오류유형 건수
        fault_types = {}
        if fault_col in df.columns:
            if primary_month and date_series is not None:
                pm_num = int(primary_month.replace("월", ""))
                mask = date_series.apply(lambda d: d is not None and d.month == pm_num)
                ft_vals = df.loc[mask, fault_col]
            else:
                ft_vals = df[fault_col]
            for ft, cnt in ft_vals.value_counts().items():
                ft_str = str(ft).strip()
                if ft_str and ft_str not in ('', 'nan', 'None', 'NaN'):
                    fault_types[ft_str] = int(cnt)

        stats[tab_name] = {
            "total": len(df),
            "total_primary": total_primary,
            "primary_month": primary_month or "",
            "top3": top3,
            "by_week": by_week,
            "by_month": by_month,
            "by_date": by_date,
            "weeks": weeks,
            "fault_by_week": fault_by_week,
            "fault_types": fault_types,
        }

    return stats


def get_date_counts_from_monthly(monthly_filename):
    """월간 파일에서 탭별 날짜별 건수 반환. {tab_name: {"M/D": count}}"""
    file_path = get_monthly_file_path(monthly_filename)
    if not os.path.exists(file_path):
        return {}
    result = {}
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        for tab_name, sheet_name in TAB_TO_RAWSHEET.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            headers, data = _read_sheet(ws)
            if not headers:
                continue
            date_col_idx = None
            for cname in ['장애접수일', '장애접수일시', '날짜']:
                if cname in headers:
                    date_col_idx = headers.index(cname)
                    break
            if date_col_idx is None:
                continue
            date_counts = {}
            for row in data:
                val = row[date_col_idx] if date_col_idx < len(row) else None
                if val is None:
                    continue
                try:
                    if isinstance(val, (date_type, datetime)):
                        d = val.date() if hasattr(val, 'date') else val
                    else:
                        from datetime import datetime as _dt
                        d = _dt.fromisoformat(str(val)[:10]).date()
                    d_str = f"{d.month}/{d.day}"
                    date_counts[d_str] = date_counts.get(d_str, 0) + 1
                except Exception:
                    continue
            result[tab_name] = date_counts
        wb.close()
    except Exception as e:
        print(f"[daily stats] {monthly_filename}: {e}")
    return result
