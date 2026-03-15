"""
analytics.py
월간 파일 통계 집계 + 분석 코멘트 생성
"""
import os
import re
import json
from datetime import date as date_type

import pandas as pd
from openpyxl import load_workbook

from excel_writer import (
    MONTHLY_FILES_DIR, TAB_TO_RAWSHEET,
    _read_sheet, get_monthly_file_path, list_monthly_files,
)

# processor 탭명 → visible 시트명
TAB_TO_VISIBLE = {
    "서울 B710":           "B710",
    "서울 B800":           "B800",
    "서울 B700":           "B700",
    "공항 B620":           "공항 B620",
    "대전 B650":           "대전 B650",
    "세종 B500":           "세종 B500",
    "제주 B400":           "제주 B400",
    "포항 B800":           "포항 B800",
    "상주,영주,예천 B400": "상주.영주.예천B400",
    "안동 B520D":          "안동B520D",
    "김해 B600":           "김해B600",
}

TAB_ORDER = [
    "서울 B800", "서울 B700", "서울 B710", "공항 B620",
    "대전 B650", "세종 B500", "제주 B400", "포항 B800",
    "상주,영주,예천 B400", "안동 B520D", "김해 B600",
]

# 지역버스 탭: 단말기접수유형/단말기현장처리 컬럼 사용 (이름에 B800이 있어도 지역버스)
REGIONAL_TABS = {
    "대전 B650", "세종 B500", "제주 B400", "포항 B800",
    "상주,영주,예천 B400", "안동 B520D", "김해 B600",
}

_CACHE_BASE = os.environ.get("DATA_DIR", os.path.dirname(__file__))
CACHE_PATH = os.path.join(_CACHE_BASE, "analytics_cache.json")  # 하위 호환용 (사용 안 함)


def _date_to_week_label(d):
    """날짜 → '3월1주' 형식 주차 레이블 (processor.get_week_label과 동일)."""
    from datetime import timedelta, date as _date, datetime as _datetime
    try:
        if isinstance(d, _datetime):
            d = d.date()
        days_to_wed = (2 - d.weekday()) % 7
        week_end_wed = d + timedelta(days=days_to_wed)
        week_start_thu = week_end_wed - timedelta(days=6)
        month = week_end_wed.month
        year = week_end_wed.year
        first_day = _date(year, month, 1)
        first_wednesday = first_day + timedelta(days=(2 - first_day.weekday()) % 7)
        first_thursday = first_wednesday - timedelta(days=6)
        week_num = (week_start_thu - first_thursday).days // 7 + 1
        return f"{month}월{week_num}주"
    except Exception:
        return ""


def _parse_ym(filename):
    m = re.search(r'(\d{4})년\s*(\d{2})월', filename)
    return (int(m.group(1)), int(m.group(2))) if m else (None, None)


def _extract_운영수량(ws):
    """visible 시트 첫 3행에서 운영수량 정수 추출. read_only/normal 모두 지원."""
    try:
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            for val in row:
                if val and "운영수량" in str(val):
                    m = re.search(r'([0-9,]+)\s*대', str(val))
                    if m:
                        return int(m.group(1).replace(',', ''))
    except Exception:
        pass
    return None


def read_운영수량_only(filename):
    """월간 파일에서 탭별 운영수량만 빠르게 읽기."""
    path = get_monthly_file_path(filename)
    if not os.path.exists(path):
        return {}
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        result = {}
        for tab_name, visible in TAB_TO_VISIBLE.items():
            if visible in wb.sheetnames:
                val = _extract_운영수량(wb[visible])
                if val:
                    result[tab_name] = val
        wb.close()
        return result
    except Exception:
        return {}


def _count_month_records(headers, rows, year, month):
    """rows 중 실제 날짜가 year/month인 행 수 반환."""
    date_pos = None
    for i, h in enumerate(headers):
        if h == "장애접수일":
            date_pos = i
            break
    if date_pos is None:
        occ = [i for i, h in enumerate(headers) if h == "장애접수일시"]
        date_pos = occ[1] if len(occ) >= 2 else (occ[0] if occ else None)

    count = 0
    for row in rows:
        if date_pos is None or date_pos >= len(row):
            count += 1
            continue
        val = row[date_pos]
        if val is None:
            continue
        try:
            d = val if isinstance(val, date_type) else pd.to_datetime(str(val)).date()
            if d.year == year and d.month == month:
                count += 1
        except Exception:
            pass
    return count


def read_monthly_stats(filename):
    """월간 파일에서 탭별 {count, 운영수량, fault_rate, by_week, by_week_top3} 반환."""
    path = get_monthly_file_path(filename)
    year, month = _parse_ym(filename)
    if not year or not os.path.exists(path):
        return {}

    result = {}
    try:
        wb = load_workbook(path, data_only=True, read_only=True)

        # 로우데이터 시트에서 건수 카운트 + 주차별 건수 + TOP3 (단일 패스)
        for tab_name in TAB_ORDER:
            raw_sheet = TAB_TO_RAWSHEET.get(tab_name)
            if not raw_sheet or raw_sheet not in wb.sheetnames:
                continue
            is_regional = tab_name in REGIONAL_TABS
            fault_col_name  = "단말기접수유형" if is_regional else "접수오류유형"
            field_col_name  = "단말기현장처리" if is_regional else "현장처리유형"
            count = 0
            by_week = {}
            by_week_faults = {}  # {week: {fault_type: count}}
            by_month_faults = {fault_col_name: {}, field_col_name: {}}  # {col: {type: count}}
            try:
                date_pos = None
                week_pos = None
                fault_pos = None
                field_pos = None
                for i, row in enumerate(wb[raw_sheet].iter_rows(values_only=True)):
                    if i == 0:
                        headers = list(row)
                        for j, h in enumerate(headers):
                            if h == "장애접수일":
                                date_pos = j
                            if h == "주차":
                                week_pos = j
                            if h == fault_col_name:
                                fault_pos = j
                            if h == field_col_name:
                                field_pos = j
                        if date_pos is None:
                            occ = [j for j, h in enumerate(headers) if h == "장애접수일시"]
                            date_pos = occ[1] if len(occ) >= 2 else (occ[0] if occ else None)
                        continue
                    if not any(v is not None for v in row):
                        continue
                    # 날짜 파싱
                    d = None
                    if date_pos is not None and date_pos < len(row) and row[date_pos] is not None:
                        try:
                            val = row[date_pos]
                            d = val if isinstance(val, date_type) else pd.to_datetime(str(val)).date()
                        except Exception:
                            pass
                    # 월 건수: 날짜 기준으로 카운트 (날짜가 해당 월인 행만)
                    if d is None:
                        if date_pos is None:
                            count += 1
                    elif d.year == year and d.month == month:
                        count += 1
                        # 월별 장애유형 집계 (접수오류유형 + 현장처리유형)
                        for pos, col in [(fault_pos, fault_col_name), (field_pos, field_col_name)]:
                            if pos is not None and pos < len(row):
                                fv = row[pos]
                                if fv not in (None, "", "nan", "None", "NaN"):
                                    fv = str(fv).strip()
                                    if fv and fv not in ("nan", "None", "NaN"):
                                        by_month_faults[col][fv] = by_month_faults[col].get(fv, 0) + 1

                    # 주차별 건수: 날짜 필터 없이 집계
                    # (예: 2월26~28일은 날짜상 2월이지만 3월1주차에 속함 → 포함해야 함)
                    wk = None
                    if week_pos is not None and week_pos < len(row) and row[week_pos]:
                        wk = str(row[week_pos])
                    elif d is not None:
                        wk = _date_to_week_label(d)
                    if wk and any(v is not None for v in row):
                        by_week[wk] = by_week.get(wk, 0) + 1
                        # TOP3 장애유형 집계
                        if fault_pos is not None and fault_pos < len(row):
                            fv = row[fault_pos]
                            if fv not in (None, "", "nan", "None", "NaN"):
                                fv = str(fv).strip()
                                if fv and fv not in ("nan", "None", "NaN"):
                                    wk_faults = by_week_faults.setdefault(wk, {})
                                    wk_faults[fv] = wk_faults.get(fv, 0) + 1
            except Exception:
                count = 0
                by_week = {}
                by_week_faults = {}
                by_month_faults = {fault_col_name: {}, field_col_name: {}}

            # 주차별 TOP3 문자열 생성
            by_week_top3 = {
                wk: "/".join(sorted(fc, key=fc.get, reverse=True)[:3])
                for wk, fc in by_week_faults.items()
            }

            result[tab_name] = {
                "year": year, "month": month,
                "count": count,
                "by_week": by_week,
                "by_week_top3": by_week_top3,
                "by_week_faults": by_week_faults,  # raw Counter (합산용)
                "by_month_faults": by_month_faults,  # {col_name: {type: count}}
                "운영수량": None,
                "fault_rate": None,
            }

        # 운영수량 (visible 시트에서)
        for tab_name in list(result.keys()):
            visible = TAB_TO_VISIBLE.get(tab_name)
            if visible and visible in wb.sheetnames:
                op = _extract_운영수량(wb[visible])
                if op:
                    result[tab_name]["운영수량"] = op
                    cnt = result[tab_name]["count"]
                    if cnt > 0:
                        result[tab_name]["fault_rate"] = round(cnt / op * 100, 2)

        wb.close()
    except Exception:
        pass

    return result


def get_all_historical_stats(force_refresh=False):
    """DB에서 전체 월별 통계 반환. DB가 비어 있으면 Excel에서 마이그레이션 후 반환."""
    import database as db

    if force_refresh or db.is_db_empty():
        # DB가 비었거나 강제 새로고침: 모든 Excel 파일 재스캔
        files = list_monthly_files()
        for fn in files:
            year, month = _parse_ym(fn)
            if not year:
                continue
            try:
                stats = read_monthly_stats(fn)
                if stats:
                    db.upsert_file_stats(fn, year, month, stats)
            except Exception as e:
                print(f"[analytics] {fn} 읽기 오류: {e}")

    return db.get_all_historical_stats()


# ─── 자동 코멘트 생성 ─────────────────────────────────────

def _fmt(n):
    return f"{n:,}"


def _pct_txt(diff, base):
    """변화율 텍스트. base=0이면 빈 문자열."""
    if not base:
        return ""
    p = round(abs(diff) / base * 100, 1)
    return f" ({p}%)" if p else ""


def _magnitude(diff, base):
    """변화 크기 형용사."""
    if not base or not diff:
        return ""
    p = abs(diff) / base * 100
    if p >= 30:
        return "대폭 "
    if p >= 10:
        return ""
    return "소폭 "


def generate_comments(all_stats, from_key=None, to_key=None):
    """확정 월간 데이터 기반 자동 분석 코멘트 생성.
    from_key: 비교 기준 월(이전), to_key: 비교 대상 월(현재).
    미지정 시 마지막 두 월로 자동 설정.
    """
    comments = []
    sorted_keys = sorted(all_stats.keys())
    if not sorted_keys:
        return comments

    # 비교할 두 월 결정
    if from_key and to_key and from_key in all_stats and to_key in all_stats:
        latest_key = to_key
        prev_key = from_key
    elif len(sorted_keys) >= 2:
        latest_key = sorted_keys[-1]
        prev_key = sorted_keys[-2]
    else:
        latest_key = sorted_keys[-1]
        prev_key = None

    latest = all_stats[latest_key]
    prev = all_stats[prev_key] if prev_key else None

    # 3개월 연속 경보용: prev_key 이전 월
    prev2 = None
    if prev_key:
        prev_idx = sorted_keys.index(prev_key)
        if prev_idx > 0:
            prev2 = all_stats[sorted_keys[prev_idx - 1]]

    lm = latest["month"]
    pm = prev["month"] if prev else "-"

    # ── 전체 요약 ──
    total_cur = sum(t.get("count", 0) for t in latest["tabs"].values())
    if prev:
        total_prv = sum(t.get("count", 0) for t in prev["tabs"].values())
        diff = total_cur - total_prv
        pct = _pct_txt(diff, total_prv)
        mag = _magnitude(diff, total_prv)
        ctype = "summary_bad" if diff > 0 else ("summary_good" if diff < 0 else "summary_same")
        if diff > 0:
            trend = f"{mag}증가하였습니다{pct}. 전체적인 장애 건수가 늘어난 만큼 주요 원인 파악이 필요합니다."
        elif diff < 0:
            trend = f"{mag}감소하였습니다{pct}. 장애 감소 추세로 개선 효과가 나타나고 있습니다."
        else:
            trend = "전월과 동일한 수준을 유지하였습니다."
        comments.append({
            "type": ctype, "tag": "전체",
            "text": f"전체 장애가 {pm}월 {_fmt(total_prv)}건에서 {lm}월 {_fmt(total_cur)}건으로 {trend}",
        })

    # ── 탭별 전월 대비 ──
    for tab in TAB_ORDER:
        cur_data = latest["tabs"].get(tab, {})
        prv_data = prev["tabs"].get(tab, {}) if prev else {}
        cur_cnt = cur_data.get("count", 0)
        prv_cnt = prv_data.get("count", 0)
        if cur_cnt == 0 and prv_cnt == 0:
            continue

        diff = cur_cnt - prv_cnt
        pct = _pct_txt(diff, prv_cnt)
        mag = _magnitude(diff, prv_cnt)
        ctype = "tab_bad" if diff > 0 else ("tab_good" if diff < 0 else "tab_same")

        if diff > 0:
            trend = f"전월 대비 {mag}{_fmt(abs(diff))}건{pct} 증가"
        elif diff < 0:
            trend = f"전월 대비 {mag}{_fmt(abs(diff))}건{pct} 감소"
        else:
            trend = "전월과 동일"

        # 장애율 변화
        rate_txt = ""
        cur_r = cur_data.get("fault_rate")
        prv_r = prv_data.get("fault_rate")
        if cur_r is not None and prv_r is not None:
            rd = round(cur_r - prv_r, 2)
            if abs(rd) >= 0.05:
                rate_arrow = "▲" if rd > 0 else "▼"
                rate_txt = f", 장애율 {prv_r}%→{cur_r}%({rate_arrow}{abs(rd)}%p)"
            else:
                rate_txt = f", 장애율 {cur_r}%"

        # 추가 코멘트 (증가/감소 이유 추측)
        remark = ""
        if diff > 0 and cur_r is not None and cur_r >= 1.0:
            remark = " 장애율이 기준(1%)을 초과하고 있어 집중 점검이 필요합니다."
        elif diff < 0 and abs(diff) >= 10:
            remark = " 꾸준한 관리로 인한 개선으로 추정됩니다."
        elif diff > 0 and tab.startswith("서울"):
            remark = " 운행 집중 구간 특성상 계절적 요인이나 노후 장비 영향을 검토해 보세요."

        text = f"{tab}: {pm}월 {_fmt(prv_cnt)}건 → {lm}월 {_fmt(cur_cnt)}건 — {trend}{rate_txt}.{remark}"
        comments.append({"type": ctype, "tag": tab, "text": text})

    # ── 3개월 연속 증가 경보 ──
    if prev and prev2:
        for tab in TAB_ORDER:
            c0 = prev2["tabs"].get(tab, {}).get("count", 0)
            c1 = prev["tabs"].get(tab, {}).get("count", 0)
            c2 = latest["tabs"].get(tab, {}).get("count", 0)
            if c0 > 0 and c1 > c0 and c2 > c1:
                comments.append({
                    "type": "alert", "tag": tab,
                    "text": (
                        f"⚠ {tab}: {prev2['month']}월 {c0}건 → {prev['month']}월 {c1}건 → {lm}월 {c2}건으로 "
                        f"3개월 연속 증가 중입니다. 원인 분석 및 현장 점검이 시급합니다."
                    ),
                })

    # ── 최고/최저 장애율 ──
    rates = [(t, d.get("fault_rate")) for t, d in latest["tabs"].items() if d.get("fault_rate")]
    if rates:
        rates.sort(key=lambda x: x[1], reverse=True)
        top_t, top_r = rates[0]
        low_t, low_r = rates[-1]
        top_txt = f"{lm}월 장애율 최고: {top_t} {top_r}%"
        if len(rates) > 1:
            top_txt += f"  |  최저: {low_t} {low_r}%"
        if top_r >= 1.0:
            top_txt += f"  — {top_t}은(는) 장애율 기준(1%)을 초과하였습니다."
        comments.append({"type": "rate_info", "tag": "장애율", "text": top_txt})

    return comments


def _is_confirmed_month(year, month):
    """확정 기간(2025년 6월 ~ 2026년 2월)인지 여부."""
    if year < 2025:
        return True
    if year == 2025 and month >= 6:
        return True
    if year == 2026 and month <= 2:
        return True
    return False


def generate_upload_comparison(new_stats, all_stats):
    """신규 업로드 데이터를 직전 확정 월과 비교하는 코멘트 생성. (메인 페이지용)
    - 업로드 데이터가 이미 확정된 월이면 빈 리스트 반환 (표시 안 함)
    - 미확정 월(예: 3월)이면 직전 확정 월과 비교
    """
    if not new_stats or not all_stats:
        return []

    sorted_keys = sorted(all_stats.keys())
    if not sorted_keys:
        return []

    # 신규 데이터의 주 월(숫자) 파악
    new_month_nums = set()
    for td in new_stats.values():
        pm = td.get("primary_month", "")
        if pm:
            try:
                new_month_nums.add(int(pm.replace("월", "")))
            except Exception:
                pass
    if not new_month_nums:
        return []
    new_month_num = max(new_month_nums)

    # all_stats에서 신규 데이터 월과 일치하는 확정 키 탐색
    matched_confirmed_key = None
    for k in sorted_keys:
        entry = all_stats[k]
        if entry.get("month") == new_month_num and _is_confirmed_month(entry.get("year", 0), entry.get("month", 0)):
            matched_confirmed_key = k
            break

    # 확정 월 업로드 → 비교 표시 안 함
    if matched_confirmed_key:
        return []

    # 미확정 월 업로드: 직전 확정 월을 찾아 비교
    prev = None
    for k in reversed(sorted_keys):
        entry = all_stats[k]
        if _is_confirmed_month(entry.get("year", 0), entry.get("month", 0)):
            prev = entry
            break
    if not prev:
        return []

    comments = []
    new_month_label = f"{new_month_num}월"
    pm_label = f"{prev['month']}월"

    # 전체 요약
    total_new = sum(td.get("total_primary", 0) for td in new_stats.values())
    total_prv = sum(t.get("count", 0) for t in prev["tabs"].values())
    diff = total_new - total_prv
    pct = _pct_txt(diff, total_prv)
    mag = _magnitude(diff, total_prv)
    ctype = "summary_bad" if diff > 0 else ("summary_good" if diff < 0 else "summary_same")
    if diff > 0:
        trend = f"{mag}증가{pct}"
    elif diff < 0:
        trend = f"{mag}감소{pct}"
    else:
        trend = "동일"
    comments.append({
        "type": ctype, "tag": "전체",
        "text": f"전체: {pm_label} {_fmt(total_prv)}건 → {new_month_label} {_fmt(total_new)}건 ({trend})",
    })

    # 탭별 비교
    for tab in TAB_ORDER:
        ns = new_stats.get(tab)
        if not ns:
            continue
        new_cnt = ns.get("total_primary", 0)
        prv_cnt = prev["tabs"].get(tab, {}).get("count", 0)
        if new_cnt == 0 and prv_cnt == 0:
            continue

        diff = new_cnt - prv_cnt
        pct = _pct_txt(diff, prv_cnt)
        mag = _magnitude(diff, prv_cnt)
        ctype = "new_bad" if diff > 0 else ("new_good" if diff < 0 else "new_same")

        if diff > 0:
            trend = f"{mag}{_fmt(abs(diff))}건{pct} 증가"
        elif diff < 0:
            trend = f"{mag}{_fmt(abs(diff))}건{pct} 감소"
        else:
            trend = "전월과 동일"

        comments.append({
            "type": ctype, "tag": tab,
            "text": f"{tab}: {pm_label} {_fmt(prv_cnt)}건 → {new_month_label} {_fmt(new_cnt)}건 — {trend}",
        })

    return comments


def generate_week_comparison(new_stats, all_stats):
    """신규 업로드 주차 데이터를 직전 주차와 비교하는 코멘트 생성.
    예: 3월1주 업로드 → 2월 마지막 주차와 비교
    """
    if not new_stats or not all_stats:
        return []

    # 신규 데이터에서 주차 목록 수집
    all_new_weeks = {}
    for tab, td in new_stats.items():
        for wk, cnt in (td.get("by_week") or {}).items():
            all_new_weeks[wk] = all_new_weeks.get(wk, 0) + cnt

    if not all_new_weeks:
        return []

    # 대표 주차: 가장 건수 많은 주차
    cur_week = max(all_new_weeks, key=lambda w: all_new_weeks[w])

    # 이전 주차 결정
    import re
    m = re.match(r'(\d+)월(\d+)주', cur_week)
    if not m:
        return []
    cur_mn, cur_wn = int(m.group(1)), int(m.group(2))

    sorted_keys = sorted(all_stats.keys())
    if cur_wn > 1:
        # 같은 달 직전 주차
        prev_week = f"{cur_mn}월{cur_wn - 1}주"
        # 해당 월 key 찾기
        prev_key = next((k for k in reversed(sorted_keys)
                         if all_stats[k]["month"] == cur_mn), None)
    else:
        # 직전 달 마지막 주차
        prev_mn = 12 if cur_mn == 1 else cur_mn - 1
        prev_key = next((k for k in reversed(sorted_keys)
                         if all_stats[k]["month"] == prev_mn), None)
        if not prev_key:
            return []
        # 직전 달의 가장 높은 주차 번호 찾기
        all_weeks_prev = set()
        for tab_data in all_stats[prev_key]["tabs"].values():
            for wk in (tab_data.get("by_week") or {}).keys():
                all_weeks_prev.add(wk)
        if not all_weeks_prev:
            return []
        # 주차 번호 기준 최대값
        def week_num(w):
            mm = re.match(r'\d+월(\d+)주', w)
            return int(mm.group(1)) if mm else 0
        prev_week = max(all_weeks_prev, key=week_num)

    if not prev_key:
        return []

    prev_data = all_stats[prev_key]["tabs"]
    comments = []

    # 전체 합계
    total_new = sum(
        (new_stats.get(tab) or {}).get("by_week", {}).get(cur_week, 0)
        for tab in TAB_ORDER
    )
    total_prv = sum(
        (prev_data.get(tab) or {}).get("by_week", {}).get(prev_week, 0)
        for tab in TAB_ORDER
    )
    diff = total_new - total_prv
    pct = _pct_txt(diff, total_prv)
    ctype = "summary_bad" if diff > 0 else ("summary_good" if diff < 0 else "summary_same")
    arrow = "▲" if diff > 0 else ("▼" if diff < 0 else "→")
    comments.append({
        "type": ctype, "tag": "전체",
        "text": f"전체: {prev_week} {_fmt(total_prv)}건 → {cur_week} {_fmt(total_new)}건 ({arrow}{_fmt(abs(diff))}건{pct})",
    })

    for tab in TAB_ORDER:
        new_cnt = (new_stats.get(tab) or {}).get("by_week", {}).get(cur_week, 0)
        prv_cnt = (prev_data.get(tab) or {}).get("by_week", {}).get(prev_week, 0)
        if new_cnt == 0 and prv_cnt == 0:
            continue

        diff = new_cnt - prv_cnt
        pct = _pct_txt(diff, prv_cnt)
        mag = _magnitude(diff, prv_cnt)
        ctype = "new_bad" if diff > 0 else ("new_good" if diff < 0 else "new_same")
        if diff > 0:
            trend = f"{mag}{_fmt(abs(diff))}건{pct} 증가"
        elif diff < 0:
            trend = f"{mag}{_fmt(abs(diff))}건{pct} 감소"
        else:
            trend = "동일"

        comments.append({
            "type": ctype, "tag": tab,
            "text": f"{tab}: {prev_week} {_fmt(prv_cnt)}건 → {cur_week} {_fmt(new_cnt)}건 — {trend}",
        })

    return comments


# ─── 스마트 비교 (신규 UI용) ─────────────────────────────────────────────────

def _cmp_rows(cur_by_tab, prev_by_tab, cur_fault_by_tab=None, prev_fault_by_tab=None):
    """탭별 현재/이전 건수 비교 rows 생성."""
    rows = []
    total_cur = sum(cur_by_tab.get(t, 0) for t in TAB_ORDER)
    total_prev = sum(prev_by_tab.get(t, 0) for t in TAB_ORDER)
    diff = total_cur - total_prev
    pct = round(diff / total_prev * 100, 1) if total_prev else None
    rows.append({"tab": "전체", "prev": total_prev, "cur": total_cur, "diff": diff, "pct": pct, "is_total": True, "fault_changes": []})
    for tab in TAB_ORDER:
        cur = cur_by_tab.get(tab, 0)
        prev = prev_by_tab.get(tab, 0)
        if cur == 0 and prev == 0:
            continue
        d = cur - prev
        p = round(d / prev * 100, 1) if prev else None
        # 오류유형 변화
        fault_changes = []
        if cur_fault_by_tab and prev_fault_by_tab:
            cur_f = cur_fault_by_tab.get(tab) or {}
            prev_f = prev_fault_by_tab.get(tab) or {}
            all_types = set(cur_f) | set(prev_f)
            for ft in sorted(all_types):
                fc = cur_f.get(ft, 0)
                fp = prev_f.get(ft, 0)
                fd = fc - fp
                if fd != 0:
                    fault_changes.append({"name": ft, "prev": fp, "cur": fc, "diff": fd})
            # diff 절댓값 내림차순 정렬
            fault_changes.sort(key=lambda x: abs(x["diff"]), reverse=True)
        rows.append({"tab": tab, "prev": prev, "cur": cur, "diff": d, "pct": p, "is_total": False, "fault_changes": fault_changes})
    return rows


def _week_cmp_block(new_stats, all_stats, all_new_weeks):
    """전주 비교 블록 생성."""
    cur_week = max(all_new_weeks, key=lambda w: all_new_weeks[w])
    m = re.match(r'(\d+)월(\d+)주', cur_week)
    if not m:
        return None
    cur_mn, cur_wn = int(m.group(1)), int(m.group(2))
    sorted_keys = sorted(all_stats.keys())
    if cur_wn > 1:
        prev_week = f"{cur_mn}월{cur_wn - 1}주"
        prev_key = next((k for k in reversed(sorted_keys) if all_stats[k]["month"] == cur_mn), None)
    else:
        prev_mn = 12 if cur_mn == 1 else cur_mn - 1
        prev_key = next((k for k in reversed(sorted_keys) if all_stats[k]["month"] == prev_mn), None)
        if not prev_key:
            return None
        all_weeks_prev = set()
        for td in all_stats[prev_key]["tabs"].values():
            for wk in (td.get("by_week") or {}).keys():
                all_weeks_prev.add(wk)
        if not all_weeks_prev:
            return None
        prev_week = max(all_weeks_prev, key=lambda w: int(re.search(r'(\d+)주', w).group(1)) if re.search(r'(\d+)주', w) else 0)
    if not prev_key:
        return None
    prev_data = all_stats[prev_key]["tabs"]
    cur_by_tab = {tab: (new_stats.get(tab) or {}).get("by_week", {}).get(cur_week, 0) for tab in TAB_ORDER}
    prev_by_tab = {tab: (prev_data.get(tab) or {}).get("by_week", {}).get(prev_week, 0) for tab in TAB_ORDER}
    # 오류유형 데이터 준비
    cur_fault_by_tab = {tab: (new_stats.get(tab) or {}).get("fault_by_week", {}).get(cur_week, {}) for tab in TAB_ORDER}
    prev_fault_by_tab = {tab: (prev_data.get(tab) or {}).get("by_week_faults", {}).get(prev_week, {}) for tab in TAB_ORDER}
    rows = _cmp_rows(cur_by_tab, prev_by_tab, cur_fault_by_tab, prev_fault_by_tab)
    if not rows:
        return None
    return {"type": "week", "label": "전주 대비", "prev_label": prev_week, "cur_label": cur_week, "rows": rows}


def generate_comparison(new_stats, all_stats):
    """스마트 비교: 데이터 범위에 따라 전일/전주 비교 생성."""
    if not new_stats or not all_stats:
        return {"comparisons": []}

    comparisons = []

    # 날짜 수집
    all_dates = set()
    for td in new_stats.values():
        for d in (td.get("by_date") or {}).keys():
            all_dates.add(d)

    # 주차 수집
    all_new_weeks = {}
    for td in new_stats.values():
        for wk, cnt in (td.get("by_week") or {}).items():
            all_new_weeks[wk] = all_new_weeks.get(wk, 0) + cnt

    n_dates = len(all_dates)
    n_weeks = len(all_new_weeks)
    is_short_range = 1 <= n_dates <= 4  # 1~4일치만 전일 비교 (5일+ = 주차 수준)
    is_monthly = n_weeks >= 3  # 3주치 이상 = 월 단위 데이터

    # ── 전월 비교 (한달치 데이터) ────────────────────────
    if is_monthly:
        new_month_nums = set()
        for td in new_stats.values():
            pm = td.get("primary_month", "")
            if pm:
                try:
                    new_month_nums.add(int(pm.replace("월", "")))
                except Exception:
                    pass
        if new_month_nums:
            new_month_num = max(new_month_nums)
            new_month_label = f"{new_month_num}월"
            sorted_keys = sorted(all_stats.keys())
            prev = None
            for k in reversed(sorted_keys):
                entry = all_stats[k]
                if entry.get("month") != new_month_num:
                    prev = entry
                    break
            if prev:
                pm_label = f"{prev['month']}월"
                cur_by_tab = {tab: (new_stats.get(tab) or {}).get("total_primary", 0) for tab in TAB_ORDER}
                prev_by_tab = {tab: (prev["tabs"].get(tab) or {}).get("count", 0) for tab in TAB_ORDER}
                cur_fault_by_tab = {tab: (new_stats.get(tab) or {}).get("fault_types", {}) for tab in TAB_ORDER}
                prev_fault_by_tab = {tab: (prev["tabs"].get(tab) or {}).get("by_month_faults", {}).get("접수오류유형",
                                    (prev["tabs"].get(tab) or {}).get("by_month_faults", {}).get("단말기접수유형", {}))
                                    for tab in TAB_ORDER}
                rows = _cmp_rows(cur_by_tab, prev_by_tab, cur_fault_by_tab, prev_fault_by_tab)
                if rows:
                    comparisons.append({
                        "type": "month",
                        "label": "전월 대비",
                        "prev_label": pm_label,
                        "cur_label": new_month_label,
                        "rows": rows,
                    })

    # ── 전일/전날 비교 ──────────────────────────────────
    if is_short_range:
        from datetime import date as _date, timedelta
        sorted_dates = sorted(all_dates, key=lambda d: (int(d.split('/')[0]), int(d.split('/')[1])))

        def _parse(d_str):
            try:
                mo, dy = d_str.split('/')
                return _date(date_type.today().year, int(mo), int(dy))
            except Exception:
                return None

        cur_date_objs = [o for o in [_parse(d) for d in sorted_dates] if o]
        if cur_date_objs:
            n_days = len(cur_date_objs)
            min_cur = min(cur_date_objs)
            prev_date_objs = list(reversed([min_cur - timedelta(days=i + 1) for i in range(n_days)]))
            prev_dates = [f"{d.month}/{d.day}" for d in prev_date_objs]

            # 월간 파일에서 전날 데이터 읽기
            primary_months = set()
            for td in new_stats.values():
                pm = td.get("primary_month", "")
                if pm:
                    try:
                        primary_months.add(int(pm.replace("월", "")))
                    except Exception:
                        pass

            monthly_daily = {}
            if primary_months:
                pm_num = max(primary_months)
                target_file = next((f for f in list_monthly_files() if f" {pm_num:02d}월" in f), None)
                if target_file:
                    from excel_writer import get_date_counts_from_monthly
                    monthly_daily = get_date_counts_from_monthly(target_file)

            if monthly_daily:
                cur_by_tab = {}
                prev_by_tab = {}
                for tab in TAB_ORDER:
                    by_date = (new_stats.get(tab) or {}).get("by_date") or {}
                    cur_by_tab[tab] = sum(by_date.get(d, 0) for d in sorted_dates)
                    day_counts = monthly_daily.get(tab) or {}
                    prev_by_tab[tab] = sum(day_counts.get(d, 0) for d in prev_dates)

                if n_days == 1:
                    cmp_label = "전일 대비"
                    prev_label = prev_dates[0]
                    cur_label = sorted_dates[0]
                else:
                    cmp_label = f"직전 {n_days}일 대비"
                    prev_label = f"{prev_dates[0]}~{prev_dates[-1]}"
                    cur_label = f"{sorted_dates[0]}~{sorted_dates[-1]}"

                rows = _cmp_rows(cur_by_tab, prev_by_tab)
                if rows:
                    comparisons.append({
                        "type": "day",
                        "label": cmp_label,
                        "prev_label": prev_label,
                        "cur_label": cur_label,
                        "rows": rows,
                    })

    # ── 전주 비교 ────────────────────────────────────────
    if all_new_weeks:
        week_block = _week_cmp_block(new_stats, all_stats, all_new_weeks)
        if week_block:
            comparisons.append(week_block)

    return {"comparisons": comparisons}
