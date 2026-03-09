from flask import Flask, render_template, request, jsonify, send_file
import os
import shutil

_BASE_DIR = os.environ.get("DATA_DIR", os.path.dirname(__file__))
_RESULTS_DIR = os.path.join(_BASE_DIR, "results")

# Railway Volume 초기화: 월간 파일이 없으면 repo의 monthly_files에서 복사
_VOLUME_MONTHLY = os.path.join(_BASE_DIR, "monthly_files")
_REPO_MONTHLY = os.path.join(os.path.dirname(__file__), "monthly_files")
if _BASE_DIR != os.path.dirname(__file__):
    os.makedirs(_VOLUME_MONTHLY, exist_ok=True)
    for _f in os.listdir(_REPO_MONTHLY):
        _dst = os.path.join(_VOLUME_MONTHLY, _f)
        if not os.path.exists(_dst):
            shutil.copy2(os.path.join(_REPO_MONTHLY, _f), _dst)
    # confirmed_weeks.json: repo에 있으면 volume으로 1회 복사 (이후엔 volume 우선)
    _repo_cw = os.path.join(os.path.dirname(__file__), "confirmed_weeks.json")
    _vol_cw = os.path.join(_BASE_DIR, "confirmed_weeks.json")
    if os.path.exists(_repo_cw) and not os.path.exists(_vol_cw):
        shutil.copy2(_repo_cw, _vol_cw)
import tempfile
import json
import pickle
from datetime import datetime
from processor import process_all_files, tabs_to_excel_bytes, build_preview
from excel_writer import list_monthly_files, insert_into_monthly, compute_web_stats, get_monthly_file_path, delete_weekly
from analytics import (
    get_all_historical_stats, generate_comments, generate_upload_comparison,
    generate_week_comparison,
    read_운영수량_only, TAB_ORDER as ANALYTICS_TAB_ORDER, _parse_ym,
)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 100 * 1024 * 1024  # 100MB

CONFIG_FILE = os.path.join(os.path.dirname(__file__), "config.json")
REFERENCE_DIR = os.path.join(os.path.dirname(__file__), "reference_files")
CONFIRMED_WEEKS_FILE = os.path.join(_BASE_DIR, "confirmed_weeks.json")

DEFAULT_CONFIG = {
    "criteria_path": "",
    "cits_path": "",
}


def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return DEFAULT_CONFIG.copy()


def save_config(cfg):
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)


def load_confirmed_weeks():
    if os.path.exists(CONFIRMED_WEEKS_FILE):
        with open(CONFIRMED_WEEKS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_confirmed_weeks(data):
    with open(CONFIRMED_WEEKS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_reference_paths():
    """reference_files/ 폴더에서 기준파일 자동 탐지. config.json 경로를 fallback으로 사용."""
    criteria, cits = None, None
    if os.path.exists(REFERENCE_DIR):
        for fn in os.listdir(REFERENCE_DIR):
            if fn.startswith('.'):
                continue
            lower = fn.lower()
            if '오류처리유형' in fn and (lower.endswith('.xls') or lower.endswith('.xlsx')):
                criteria = os.path.join(REFERENCE_DIR, fn)
            elif 'cits' in lower and (lower.endswith('.xls') or lower.endswith('.xlsx')):
                cits = os.path.join(REFERENCE_DIR, fn)
    cfg = load_config()
    criteria = criteria or cfg.get("criteria_path", "")
    cits = cits or cfg.get("cits_path", "")
    return criteria, cits


@app.route("/")
def index():
    cfg = load_config()
    return render_template("index.html", config=cfg)


@app.route("/api/config", methods=["GET"])
def get_config():
    return jsonify(load_config())


@app.route("/api/reference-files")
def get_reference_files():
    """현재 서버에 저장된 기준파일 목록 반환."""
    criteria, cits = get_reference_paths()
    return jsonify({
        "criteria": os.path.basename(criteria) if criteria and os.path.exists(criteria) else None,
        "cits": os.path.basename(cits) if cits and os.path.exists(cits) else None,
    })


@app.route("/api/upload-reference", methods=["POST"])
def upload_reference():
    """기준파일 서버 업로드 (reference_files/ 폴더에 저장)."""
    file_type = request.form.get("type")  # "criteria" or "cits"
    f = request.files.get("file")
    if not f or f.filename == "":
        return jsonify({"error": "파일 없음"}), 400
    if file_type not in ("criteria", "cits"):
        return jsonify({"error": "type 파라미터 오류"}), 400

    os.makedirs(REFERENCE_DIR, exist_ok=True)
    ext = os.path.splitext(f.filename)[1]
    save_name = f"criteria{ext}" if file_type == "criteria" else f"cits{ext}"
    save_path = os.path.join(REFERENCE_DIR, save_name)
    f.save(save_path)
    return jsonify({"ok": True, "filename": f.filename, "saved_as": save_name})


@app.route("/api/config", methods=["POST"])
def update_config():
    data = request.json
    cfg = load_config()
    if "criteria_path" in data:
        cfg["criteria_path"] = data["criteria_path"]
    if "cits_path" in data:
        cfg["cits_path"] = data["cits_path"]
    save_config(cfg)
    return jsonify({"ok": True})


def _is_confirmed(filename):
    """25년6월~26년2월 파일은 확정(잠금) 상태."""
    import re
    m = re.search(r'(\d{4})년\s*(\d{2})월', filename)
    if not m:
        return False
    year, month = int(m.group(1)), int(m.group(2))
    if year < 2025:
        return True
    if year == 2025 and month >= 6:
        return True
    if year == 2026 and month <= 2:
        return True
    return False


@app.route("/api/monthly-files")
def get_monthly_files():
    files = list_monthly_files()
    return jsonify({
        "files": [
            {"name": f, "confirmed": _is_confirmed(f)}
            for f in files
        ]
    })


@app.route("/api/process", methods=["POST"])
def process():
    cfg = load_config()

    files = request.files.getlist("files")
    if not files or all(f.filename == "" for f in files):
        return jsonify({"error": "파일이 없습니다."}), 400

    auto_criteria, auto_cits = get_reference_paths()
    criteria_path = request.form.get("criteria_path") or auto_criteria
    cits_path = request.form.get("cits_path") or auto_cits

    if not criteria_path or not os.path.exists(criteria_path):
        return jsonify({"error": "오류처리유형 기준 파일 경로를 확인해주세요. (⚙ 설정)"}), 400
    if not cits_path or not os.path.exists(cits_path):
        return jsonify({"error": "CITS 기준 파일 경로를 확인해주세요. (⚙ 설정)"}), 400

    tmp_paths = []
    try:
        for f in files:
            if f.filename == "":
                continue
            suffix = os.path.splitext(f.filename)[1]
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                f.save(tmp.name)
                tmp_paths.append(tmp.name)

        final_tabs, summary, meta = process_all_files(tmp_paths, criteria_path, cits_path)

        if not final_tabs:
            return jsonify({"error": "처리된 데이터가 없습니다. 파일 형식을 확인해주세요."}), 400

        excel_bytes = tabs_to_excel_bytes(final_tabs)

        result_dir = _RESULTS_DIR
        os.makedirs(result_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        result_filename = f"장애목록_처리결과_{timestamp}.xlsx"
        result_path = os.path.join(result_dir, result_filename)
        with open(result_path, "wb") as f:
            f.write(excel_bytes)

        # 처리 결과 저장 (월간 삽입에서 재사용)
        pkl_path = result_path + ".pkl"
        with open(pkl_path, "wb") as f:
            pickle.dump(final_tabs, f)

        previews = build_preview(final_tabs)
        stats = compute_web_stats(final_tabs)

        # 주차 목록 추출
        all_weeks = set()
        for tab_stats in stats.values():
            all_weeks.update(tab_stats.get("weeks", []))

        # 주 대상 월 감지 → 해당 월간 파일에서 운영수량 읽어 장애율 추가
        primary_months = {ts["primary_month"] for ts in stats.values() if ts.get("primary_month")}
        운영수량_map = {}
        if primary_months:
            try:
                pm_num = int(sorted(primary_months, key=lambda x: int(x.replace("월", "")))[-1].replace("월", ""))
                target_file = next(
                    (f for f in list_monthly_files() if f" {pm_num:02d}월" in f),
                    None
                )
                if target_file:
                    운영수량_map = read_운영수량_only(target_file)
            except Exception:
                pass

        for tab_name, ts in stats.items():
            op = 운영수량_map.get(tab_name)
            if op and op > 0:
                cnt = ts.get("total_primary", ts.get("total", 0))
                ts["운영수량"] = op
                ts["fault_rate"] = round(cnt / op * 100, 2)
            else:
                ts["운영수량"] = None
                ts["fault_rate"] = None

        # 최신 stats 저장 (분석 페이지 신규 vs 전월 비교용)
        try:
            latest_stats_path = os.path.join(result_dir, "latest_stats.json")
            with open(latest_stats_path, "w", encoding="utf-8") as f:
                json.dump(stats, f, ensure_ascii=False, default=str)
        except Exception:
            pass

        # 신규 업로드 vs 직전 확정 월/주차 비교 코멘트
        upload_comments = []
        week_comments = []
        try:
            hist = get_all_historical_stats()
            upload_comments = generate_upload_comparison(stats, hist)
            week_comments = generate_week_comparison(stats, hist)
        except Exception:
            pass

        return jsonify({
            "ok": True,
            "summary": summary,
            "filename": result_filename,
            "previews": previews,
            "stats": stats,
            "weeks": sorted(all_weeks),
            "upload_comments": upload_comments,
            "week_comments": week_comments,
            "file_types": {
                os.path.basename(p): m.get("type", "unknown")
                for p, m in zip(tmp_paths, meta.values())
            },
        })

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detail": traceback.format_exc()}), 500
    finally:
        for p in tmp_paths:
            try:
                os.unlink(p)
            except Exception:
                pass


@app.route("/api/insert-monthly", methods=["POST"])
def insert_monthly():
    data = request.json
    result_key = data.get("result_key")
    monthly_filename = data.get("monthly_filename")
    mode = data.get("mode", "daily")
    week_label = data.get("week_label", "")

    if not result_key or not monthly_filename:
        return jsonify({"error": "result_key 또는 monthly_filename 누락"}), 400

    if _is_confirmed(monthly_filename):
        return jsonify({"error": f"확정 완료된 파일입니다. 삽입이 불가합니다.\n({monthly_filename})"}), 403

    # 주차 확정 여부 체크
    if mode == "weekly" and week_label:
        cw = load_confirmed_weeks()
        if week_label in cw.get(monthly_filename, []):
            return jsonify({"error": f"'{week_label}'은 확정된 주차입니다. 삽입이 불가합니다."}), 403

    result_dir = _RESULTS_DIR
    pkl_path = os.path.join(result_dir, result_key + ".pkl")

    if not os.path.exists(pkl_path):
        return jsonify({"error": "처리된 데이터가 없습니다. 먼저 처리를 실행해주세요."}), 400

    try:
        import gc
        with open(pkl_path, "rb") as f:
            final_tabs = pickle.load(f)
        gc.collect()

        report = insert_into_monthly(
            final_tabs,
            monthly_filename,
            mode=mode,
            week_label=week_label if mode == "weekly" else None,
        )

        if "error" in report:
            return jsonify({"error": report["error"]}), 400

        return jsonify({"ok": True, "report": report, "filename": monthly_filename})

    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detail": traceback.format_exc()}), 500


@app.route("/api/result-data/<result_key>")
def get_result_data(result_key):
    """처리된 탭 데이터를 JSON으로 반환 (클라이언트 사이드 삽입용)."""
    import gc
    result_dir = _RESULTS_DIR
    pkl_path = os.path.join(result_dir, result_key + ".pkl")
    if not os.path.exists(pkl_path):
        return jsonify({"error": "처리된 데이터가 없습니다. 다시 처리를 실행해주세요."}), 404
    try:
        with open(pkl_path, "rb") as f:
            final_tabs = pickle.load(f)
        tabs_json = {}
        for tab_name, df in final_tabs.items():
            if df is None or df.empty:
                continue
            columns = list(df.columns)
            rows = []
            for vals in df.itertuples(index=False, name=None):
                r = []
                for v in vals:
                    if v is None:
                        r.append(None)
                    elif isinstance(v, float) and v != v:
                        r.append(None)
                    elif hasattr(v, "strftime"):
                        r.append(v.strftime("%Y-%m-%d"))
                    elif isinstance(v, (int, float, str, bool)):
                        r.append(v)
                    else:
                        r.append(str(v))
                rows.append(r)
            tabs_json[tab_name] = {"columns": columns, "rows": rows}
        del final_tabs
        gc.collect()
        return jsonify({"ok": True, "tabs": tabs_json})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detail": traceback.format_exc()}), 500


@app.route("/api/confirmed-weeks")
def get_confirmed_weeks():
    return jsonify(load_confirmed_weeks())


@app.route("/api/confirm-week", methods=["POST"])
def confirm_week():
    data = request.json
    filename = data.get("monthly_filename")
    week = data.get("week_label")
    if not filename or not week:
        return jsonify({"error": "파라미터 누락"}), 400
    cw = load_confirmed_weeks()
    cw.setdefault(filename, [])
    if week not in cw[filename]:
        cw[filename].append(week)
    save_confirmed_weeks(cw)
    return jsonify({"ok": True})


@app.route("/api/unconfirm-week", methods=["POST"])
def unconfirm_week():
    data = request.json
    filename = data.get("monthly_filename")
    week = data.get("week_label")
    if not filename or not week:
        return jsonify({"error": "파라미터 누락"}), 400
    cw = load_confirmed_weeks()
    if filename in cw and week in cw[filename]:
        cw[filename].remove(week)
    save_confirmed_weeks(cw)
    return jsonify({"ok": True})


@app.route("/api/delete-weekly", methods=["POST"])
def delete_weekly_route():
    data = request.json
    filename = data.get("monthly_filename")
    week = data.get("week_label")
    if not filename or not week:
        return jsonify({"error": "파라미터 누락"}), 400
    cw = load_confirmed_weeks()
    if week in cw.get(filename, []):
        return jsonify({"error": f"'{week}'은 확정된 주차입니다. 삭제하려면 먼저 확정을 취소하세요."}), 403
    try:
        report = delete_weekly(filename, week)
        if "error" in report:
            return jsonify({"error": report["error"]}), 400
        return jsonify({"ok": True, "report": report})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detail": traceback.format_exc()}), 500


@app.route("/api/confirmed-weeks-stats")
def confirmed_weeks_stats():
    """확정된 주차별 탭별 건수 반환."""
    try:
        cw = load_confirmed_weeks()
        if not cw:
            return jsonify({"ok": True, "weeks": []})

        from excel_writer import get_monthly_file_path, TAB_TO_RAWSHEET
        from openpyxl import load_workbook

        result = []
        for monthly_filename, week_list in sorted(cw.items()):
            file_path = get_monthly_file_path(monthly_filename)
            if not os.path.exists(file_path):
                continue
            try:
                wb = load_workbook(file_path, read_only=True, data_only=True)
                for week_label in sorted(week_list):
                    tab_counts = {}
                    for tab_name, sheet_name in TAB_TO_RAWSHEET.items():
                        if sheet_name not in wb.sheetnames:
                            continue
                        ws = wb[sheet_name]
                        rows = list(ws.values)
                        if not rows:
                            continue
                        headers = list(rows[0])
                        week_col = next(
                            (i for i, h in enumerate(headers) if h and str(h).strip() == "주차"),
                            None,
                        )
                        if week_col is None:
                            continue
                        count = sum(1 for row in rows[1:] if len(row) > week_col and row[week_col] == week_label)
                        if count > 0:
                            tab_counts[tab_name] = count
                    result.append({
                        "week_label": week_label,
                        "monthly_filename": monthly_filename,
                        "tab_counts": tab_counts,
                        "total": sum(tab_counts.values()),
                    })
                wb.close()
            except Exception:
                continue

        return jsonify({"ok": True, "weeks": result})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detail": traceback.format_exc()}), 500


@app.route("/analysis")
def analysis_page():
    return render_template("analysis.html")


@app.route("/api/analysis-data")
def analysis_data():
    try:
        all_stats = get_all_historical_stats()

        comments = generate_comments(all_stats)

        # 차트용 데이터 변환
        sorted_keys = sorted(all_stats.keys())
        labels = [all_stats[k]["label"] for k in sorted_keys]

        chart_datasets = {}
        for tab in ANALYTICS_TAB_ORDER:
            counts = []
            rates = []
            for k in sorted_keys:
                tab_data = all_stats[k]["tabs"].get(tab, {})
                counts.append(tab_data.get("count", 0))
                rates.append(tab_data.get("fault_rate"))
            chart_datasets[tab] = {"counts": counts, "rates": rates}

        # 각 월별 확정 여부 플래그 추가
        for k, v in all_stats.items():
            v["confirmed"] = _is_confirmed(v.get("filename", ""))

        return jsonify({
            "ok": True,
            "labels": labels,
            "chart_datasets": chart_datasets,
            "all_stats": all_stats,
            "comments": comments,
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detail": traceback.format_exc()}), 500


@app.route("/api/analysis-comments")
def analysis_comments():
    """선택한 기간(from/to)에 대한 코멘트 생성."""
    try:
        from_key = request.args.get("from")
        to_key = request.args.get("to")

        all_stats = get_all_historical_stats()
        comments = generate_comments(all_stats, from_key=from_key, to_key=to_key)
        return jsonify({"ok": True, "comments": comments})
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "detail": traceback.format_exc()}), 500


@app.route("/api/download/<filename>")
def download(filename):
    result_dir = _RESULTS_DIR
    path = os.path.join(result_dir, filename)
    if not os.path.exists(path):
        return jsonify({"error": "파일 없음"}), 404
    return send_file(path, as_attachment=True, download_name=filename)


@app.route("/api/download-monthly/<filename>")
def download_monthly(filename):
    file_path = get_monthly_file_path(filename)
    if not os.path.exists(file_path):
        return jsonify({"error": "파일 없음"}), 404
    return send_file(file_path, as_attachment=True, download_name=filename)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "true").lower() == "true"
    app.run(host="0.0.0.0", port=port, debug=debug)
